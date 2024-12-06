import os
import traceback
import openpyxl
import xmltodict
import glob
import pandas as pd
from sys import exit
from datetime import datetime
import xml.etree.ElementTree as ET


# codigo principal
def main():
    # obtener subdirectorios
    cur_dir = os.getcwd()
    pattern = cur_dir + '/**/*.xml'
    files = glob.glob(pattern, recursive=True)

    facturas = []

    for file in files:
        # archivo en proceso
        print(f"Procesando {file}")

        # convertir XML a diccionario
        with open(file, encoding='utf8') as fd:
            doc = xmltodict.parse(fd.read())

        """tree = ET.parse(file)
        xml_data = tree.getroot()
        #here you can change the encoding type to be able to set it to the one you need
        xmlstr = ET.tostring(xml_data, encoding='utf-8', method='xml')

        doc = dict(xmltodict.parse(xmlstr))
        """

        
        factura = {}

        factura['emisor'] = doc['cfdi:Comprobante']['cfdi:Emisor']['@Rfc'] + ' - ' + doc['cfdi:Comprobante']['cfdi:Emisor']['@Nombre']
        factura['uuid'] = doc['cfdi:Comprobante']['cfdi:Complemento']['tfd:TimbreFiscalDigital']['@UUID']
        factura['serie-folio'] = f"{doc['cfdi:Comprobante']['@Serie']}-" + \
                                f"{doc['cfdi:Comprobante']['@Folio']}"
        # convertir string en fecha
        factura['fecha'] = datetime.strptime(doc['cfdi:Comprobante']['@Fecha'], '%Y-%m-%dT%H:%M:%S')

        factura['Cant Magna'] = 0
        factura['Cant Premium'] = 0
        factura['Cant Diesel'] = 0
        factura['Cant Otro'] = 0
        factura['Cant Total'] = 0

        factura['Importe Magna'] = 0
        factura['Importe Premium'] = 0
        factura['Importe Diesel'] = 0
        factura['Importe Otro'] = 0
        factura['SubTotal'] = 0

        factura['Importe Total'] = float(doc['cfdi:Comprobante']['@Total'])
        factura['Archivo'] = file

        for concepto in doc['cfdi:Comprobante']['cfdi:Conceptos']['cfdi:Concepto']:
            clave = concepto['@ClaveProdServ']
            if clave == '15101514':
                producto = 'Magna'
            elif clave == '15101515':
                producto = 'Premium'
            elif clave == '15101505':
                producto = 'Diesel'
            else:
                producto = 'Otro'

            factura[f"Cant {producto}"] += float(concepto['@Cantidad'])
            factura["Cant Total"] += float(concepto['@Cantidad'])
            factura[f"Importe {producto}"] += float(concepto['@Importe'])

        factura['SubTotal'] = float(doc['cfdi:Comprobante']['@SubTotal'])

        # agregar factura a facturas
        facturas.append(factura)

    if len(facturas) == 0:
        raise Exception('No se encontraron facturas')

    print(f"Se encontraron {len(facturas)} facturas")

    """
    # exportar CSV
    with open('facturas.csv', 'w', newline='') as fd:
        writer = csv.DictWriter(fd, fieldnames=facturas[0].keys())
        writer.writeheader()
        for factura in facturas:
            writer.writerow(factura)
    print('facturas.csv exportado')
    """

    # Convert dictionary to DataFrame
    df = pd.DataFrame(facturas)
    # agregar una linea de encabezado
    df.head

    # Export DataFrame to Excel
    writer = pd.ExcelWriter('facturas.xlsx', engine='openpyxl')
    df.to_excel(writer, sheet_name='Facturas', index=False, startrow=1)

    # ajustar ancho de columnas
    worksheet = writer.sheets['Facturas']

    worksheet['A1'] = 'RFC y Nombre'
    worksheet.merge_cells('A1:A2')
    worksheet['B1'] = 'Factura'
    worksheet.merge_cells('B1:B2')
    worksheet['C1'] = 'Serie y Folio'
    worksheet.merge_cells('C1:C2')
    worksheet['D1'] = 'Fecha'
    worksheet.merge_cells('D1:D2')

    # linea de borde
    border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                    right=openpyxl.styles.Side(style='thin'),
                                    top=openpyxl.styles.Side(style='thin'),
                                    bottom=openpyxl.styles.Side(style='thin'))

    worksheet['F1'] = 'Cantidades'
    worksheet['F1'].alignment = openpyxl.styles.Alignment(horizontal='center')
    worksheet.merge_cells('F1:I1')
    worksheet['F1'].border = border

    worksheet['J1'] = 'Importes'
    worksheet['J1'].alignment = openpyxl.styles.Alignment(horizontal='center')
    worksheet.merge_cells('J1:N1')
    worksheet['J1'].border = border
    worksheet['N1'].border = border


    worksheet['E2'] = 'Magna'
    worksheet['F2'] = 'Premium'
    worksheet['G2'] = 'Diesel'
    worksheet['H2'] = 'Otro'
    worksheet['I2'] = 'Total Cantidad'

    worksheet['J2'] = 'Magna'
    worksheet['K2'] = 'Premium'
    worksheet['L2'] = 'Diesel'
    worksheet['M2'] = 'Otro'
    worksheet['N2'] = 'Subtotal'

    worksheet['O1'] = 'Importe Total'
    worksheet.merge_cells('O1:O2')

    worksheet['P1'] = 'Archivo'
    worksheet.merge_cells('P1:P2')

    for column in df:
        column_length = max(df[column].astype(str).map(len).max(), len(column))+2
        col_idx = df.columns.get_loc(column)
        cell = worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)]
        cell.width = column_length

    letras = ''.join(chr(i) for i in range(ord('D'), ord('O') + 1))
    for column in letras:
        for row in range(2, df.shape[0] + 3):
            # dar formato de numero a la celda
            worksheet[f"{column}{row}"].number_format = 'dd/mm/yyyy' if column == 'D' else '#,##0.00'

    writer.close()

    print('facturas.xlsx exportado')


try:
    main()
except Exception as e:
    # mostrar error completo
    print('Error:')
    print(e)
    print('Traceback:')
    print(''.join(traceback.format_tb(e.__traceback__)))

#pausar
input('Presione Enter para continuar...')
