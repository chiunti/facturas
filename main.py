import openpyxl
import xmltodict
import csv
import glob
import pandas as pd

# recursive print
def print_dict(d, indent=2):
    for key, value in d.items():
        print(' ' * indent + str(key))
        if isinstance(value, dict):
            print_dict(value, indent + 1)
        else:
            pass
            # print('\t' * (indent + 1) + str(value))


files = glob.glob('data/*.xml')

facturas = []

for file in files:
    with open(file) as fd:
        doc = xmltodict.parse(fd.read())
    
    factura = {}

    factura['emisor'] = doc['cfdi:Comprobante']['cfdi:Emisor']['@Rfc'] + ' - ' + doc['cfdi:Comprobante']['cfdi:Emisor']['@Nombre']
    factura['uuid'] = doc['cfdi:Comprobante']['cfdi:Complemento']['tfd:TimbreFiscalDigital']['@UUID']
    factura['serie-folio'] = f"{doc['cfdi:Comprobante']['@Serie']}-" + \
                             f"{doc['cfdi:Comprobante']['@Folio']}"

    factura['Cant Magna'] = 0
    factura['Cant Premium'] = 0
    factura['Cant Diesel'] = 0
    factura['Cant Otro'] = 0
    factura['Cant Total'] = 0

    factura['Importe Magna'] = 0
    factura['Importe Premium'] = 0
    factura['Importe Diesel'] = 0
    factura['Importe Otro'] = 0
    factura['Importe Total'] = 0

    for concepto in doc['cfdi:Comprobante']['cfdi:Conceptos']['cfdi:Concepto']:
        clave = concepto['@ClaveProdServ']
        if clave == '15101514':
            producto = 'Magna'
        elif clave == '15101515':
            producto = 'Premium'
        elif clave == '15101505':
            producto = 'Diesel'
        else:
            producto = 'Otro'

        factura[f"Cant {producto}"] += float(concepto['@Cantidad'])
        factura["Cant Total"] += float(concepto['@Cantidad'])
        factura[f"Importe {producto}"] += float(concepto['@Importe'])

    factura['Importe Total'] = float(doc['cfdi:Comprobante']['@SubTotal'])

    # agregar factura a facturas
    facturas.append(factura)

if len(facturas) == 0:
    print('No se encontraron facturas')
    exit()

print(f"Se encontraron {len(facturas)} facturas")

# exportar CSV
with open('facturas.csv', 'w', newline='') as fd:
    writer = csv.DictWriter(fd, fieldnames=facturas[0].keys())
    writer.writeheader()
    for factura in facturas:
        writer.writerow(factura)

print('facturas.csv exportado')

# Convert dictionary to DataFrame
df = pd.DataFrame(facturas)
# agregar una linea de encabezado
df.head

# Export DataFrame to Excel
writer = pd.ExcelWriter('facturas.xlsx', engine='openpyxl')
df.to_excel(writer, sheet_name='Facturas', index=False, startrow=1)

# ajustar ancho de columnas
workbook = writer.book
worksheet = writer.sheets['Facturas']

worksheet['A1'] = 'RFC y Nombre'
worksheet.merge_cells('A1:A2')
worksheet['B1'] = 'Factura'
worksheet.merge_cells('B1:B2')
worksheet['C1'] = 'Serie y Folio'
worksheet.merge_cells('C1:C2')

# linea de borde
border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                right=openpyxl.styles.Side(style='thin'),
                                top=openpyxl.styles.Side(style='thin'),
                                bottom=openpyxl.styles.Side(style='thin'))

worksheet['D1'] = 'Cantidades'
worksheet['D1'].alignment = openpyxl.styles.Alignment(horizontal='center')
worksheet.merge_cells('D1:H1')
worksheet['D1'].border = border

worksheet['I1'] = 'Importes'
worksheet['I1'].alignment = openpyxl.styles.Alignment(horizontal='center')
worksheet.merge_cells('I1:M1')
worksheet['I1'].border = border
worksheet['M1'].border = border


worksheet['D2'] = 'Magna'
worksheet['E2'] = 'Premium'
worksheet['F2'] = 'Diesel'
worksheet['G2'] = 'Otro'
worksheet['H2'] = 'Total Cantidad'

worksheet['I2'] = 'Magna'
worksheet['J2'] = 'Premium'
worksheet['K2'] = 'Diesel'
worksheet['L2'] = 'Otro'
worksheet['M2'] = 'Total Importe'

for column in df:
    column_length = max(df[column].astype(str).map(len).max(), len(column))
    col_idx = df.columns.get_loc(column)
    cell = worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)]
    cell.width = column_length
    if col_idx > 1:
        cell.number_format = '#,##0.00'

writer.close()

print('facturas.xlsx exportado')